from __future__ import annotations

import json
import re
import time
from pathlib import Path
from typing import Any, Dict, List, Tuple, Optional

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data" / "dataGet_api"
SURF_PATH = BASE_DIR / "data" / "currency_kinds" / "surf_pairs.json"
RESULT_DIR = BASE_DIR / "result"
RESULT_DIR.mkdir(parents=True, exist_ok=True)
RESULT_HTML_DIR = RESULT_DIR / "html"
RESULT_HTML_DIR.mkdir(parents=True, exist_ok=True)

EX_ORDER = ["binance", "weex", "mexc", "bybit", "surf"]  # 按样表顺序，增加 SURF

# 读取 surf 目标（只取 USDT）
def load_targets() -> List[str]:
    data = json.loads(SURF_PATH.read_text(encoding="utf-8"))
    outs: List[str] = []
    for p in data.get("pairs", []):
        base = str(p.get("base") or "").upper().strip()
        quote = str(p.get("quote") or "").upper().strip()
        if not base or quote != "USDT":
            continue
        outs.append(f"{base}{quote}")
    return sorted(set(outs))


# 加载四家 selected 文件

def load_binance() -> Dict[str, List[Dict[str, Any]]]:
    p = DATA_DIR / "binance" / "binance_selected.json"
    if not p.exists():
        return {}
    items = json.loads(p.read_text(encoding="utf-8"))
    # binance_selected 是列表，元素结构：{symbol, riskBrackets:[{bracketMaintenanceMarginRate, bracketNotionalCap, maxOpenPosLeverage, ...}]}
    out: Dict[str, List[Dict[str, Any]]] = {}
    if isinstance(items, list):
        for it in items:
            if not isinstance(it, dict):
                continue
            sym = str(it.get("symbol") or it.get("s") or it.get("pair") or "").strip().upper()
            if not sym:
                continue
            tiers = []
            rbs = it.get("riskBrackets") or it.get("brackets") or []
            if isinstance(rbs, list):
                for rb in rbs:
                    if not isinstance(rb, dict):
                        continue
                    lev = rb.get("maxOpenPosLeverage") or rb.get("initialLeverage")
                    notional = rb.get("bracketNotionalCap") or rb.get("notionalCap")
                    mmr = rb.get("bracketMaintenanceMarginRate") or rb.get("maintMarginRatio")
                    tiers.append({
                        "mlev": lev,
                        "notional_usdt": notional,
                        "mmr": mmr,
                    })
            out[sym] = tiers
    return out


def load_surf() -> Dict[str, List[Dict[str, Any]]]:
    """读取 SURF 限额聚合结果，将单条限额映射为统一的 tiers 结构。
    输入文件：data/dataGet_api/surf/surf_limits.json
    结构：{"items": [{symbol, pair_id, pair_name, max_leverage, max_order_size, max_mmr, ...}, ...]}
    输出：{ "ETHUSDT": [ {mlev, notional_usdt, mmr} ] }
    """
    p = DATA_DIR / "surf" / "surf_limits.json"
    if not p.exists():
        return {}
    data = json.loads(p.read_text(encoding="utf-8"))
    items = data.get("items") if isinstance(data, dict) else None
    out: Dict[str, List[Dict[str, Any]]] = {}
    if isinstance(items, list):
        for it in items:
            if not isinstance(it, dict):
                continue
            base = str(it.get("symbol") or "").strip().upper()
            if not base:
                continue
            sym = f"{base}USDT"
            mlev = it.get("max_leverage")
            notional = it.get("max_order_size")  # 字符串数值，保持一致再统一 parse
            mmr = it.get("max_mmr")              # 小数，如 0.01
            out[sym] = [{
                "mlev": mlev,
                "notional_usdt": notional,
                "mmr": mmr,
            }]
    return out

def load_bybit() -> Dict[str, List[Dict[str, Any]]]:
    p = DATA_DIR / "bybit" / "bybit_selected.json"
    if not p.exists():
        return {}
    data = json.loads(p.read_text(encoding="utf-8"))  # 结构：{symbol: [ {maximumLever, storingLocationValue, maintenanceMarginRate, ...}, ... ]}
    out: Dict[str, List[Dict[str, Any]]] = {}
    if isinstance(data, dict):
        for sym, tiers in data.items():
            if not isinstance(tiers, list):
                continue
            out[str(sym).upper()] = tiers
    return out


def load_mexc() -> Dict[str, List[Dict[str, Any]]]:
    p = DATA_DIR / "mexc" / "mexc_selected.json"
    if not p.exists():
        return {}
    data = json.loads(p.read_text(encoding="utf-8"))  # 结构：{symbol: [ {mlev, notional_usdt, mmr, ...}, ... ]}
    out: Dict[str, List[Dict[str, Any]]] = {}
    if isinstance(data, dict):
        for sym, tiers in data.items():
            if not isinstance(tiers, list):
                continue
            out[str(sym).upper()] = tiers
    return out


def load_weex() -> Dict[str, List[Dict[str, Any]]]:
    p = DATA_DIR / "weex" / "weex_selected.json"
    if not p.exists():
        return {}
    data = json.loads(p.read_text(encoding="utf-8"))  # 结构：{symbol: [ {lv, range, mlev, mmr}, ... ]}
    out: Dict[str, List[Dict[str, Any]]] = {}
    if isinstance(data, dict):
        for sym, tiers in data.items():
            if not isinstance(tiers, list):
                continue
            out[str(sym).upper()] = tiers
    return out


# 规范化工具
NUM_RE = re.compile(r"[-+]?[0-9]*\.?[0-9]+")


def parse_number(v: Any) -> Optional[float]:
    try:
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip()
        m = NUM_RE.search(s)
        if not m:
            return None
        return float(m.group(0))
    except Exception:
        return None


def to_percent_str(v: Any) -> str:
    num = parse_number(v)
    if num is None:
        return ""
    # 若已经是百分数形式（如 "1.0%"），直接返回
    if isinstance(v, str) and v.strip().endswith('%'):
        try:
            # 统一两位小数
            return f"{float(v.strip().rstrip('%')):.2f}%"
        except Exception:
            return v
    return f"{num * 100:.2f}%"


def to_leverage_str(v: Any) -> str:
    """将不同形式的杠杆值转为 '<value>X'，保留小数，不强制取整。
    优先使用原始字符串中的数字表示；若为数值则按小数去除尾随0显示。
    """
    if v is None:
        return ""
    # 若原本是字符串，尽量保留其小数表达
    if isinstance(v, str):
        s = v.strip()
        # 如果已包含 x/X，统一为大写X并去空格
        if 'x' in s.lower():
            s_clean = s.lower().replace('x', 'X')
            s_clean = s_clean.replace(' ', '')
            # 统一成 '<number>X' 形式
            m = NUM_RE.search(s_clean)
            if m:
                num_txt = m.group(0)
                return f"{num_txt}X"
            return s_clean
        # 不含 X，则提取数字文本
        m = NUM_RE.search(s)
        if m:
            num_txt = m.group(0)
            return f"{num_txt}X"
        return ""
    # 数值：格式化，保留小数但去掉多余0
    try:
        f = float(v)
        if f.is_integer():
            return f"{int(f)}X"
        # 保留最多6位小数，去尾随0
        s = f"{f:.6f}".rstrip('0').rstrip('.')
        return f"{s}X"
    except Exception:
        num = parse_number(v)
        if num is None:
            return ""
        if float(num).is_integer():
            return f"{int(num)}X"
        s = f"{float(num):.6f}".rstrip('0').rstrip('.')
        return f"{s}X"


def weex_range_upper(s: str) -> Optional[float]:
    if not isinstance(s, str):
        return None
    if "~" in s:
        try:
            right = s.split("~", 1)[1]
            return parse_number(right)
        except Exception:
            return None
    return parse_number(s)


# 构造单个交易所的行块

def build_rows_for_exchange(ex: str, sym: str, sources: Dict[str, Dict[str, List[Dict[str, Any]]]]) -> List[List[Any]]:
    tiers: List[Dict[str, Any]] = []
    if ex == "binance":
        tiers = sources[ex].get(sym, [])
        # 按杠杆从小到大
        tiers = sorted(tiers, key=lambda x: parse_number(x.get("mlev")) or 0)
        rows: List[List[Any]] = []
        for i, t in enumerate(tiers):
            ex_name = "BINANCE" if i == 0 else ""
            lev = to_leverage_str(t.get("mlev"))
            notional = parse_number(t.get("notional_usdt")) or parse_number(t.get("bracketNotionalCap"))
            mmr = to_percent_str(t.get("mmr"))
            rows.append([ex_name, lev, notional, mmr])
        return rows or [["BINANCE", "", "", ""]]
    elif ex == "bybit":
        tiers = sources[ex].get(sym, [])
        tiers = sorted(tiers, key=lambda x: parse_number(x.get("maximumLever")) or 0)
        rows = []
        for i, t in enumerate(tiers):
            ex_name = "BYBIT" if i == 0 else ""
            lev = to_leverage_str(t.get("maximumLever"))
            notional = parse_number(t.get("storingLocationValue"))
            mmr = to_percent_str(t.get("maintenanceMarginRate"))
            rows.append([ex_name, lev, notional, mmr])
        return rows or [["BYBIT", "", "", ""]]
    elif ex == "mexc":
        tiers = sources[ex].get(sym, [])
        tiers = sorted(tiers, key=lambda x: parse_number(x.get("mlev")) or 0)
        rows = []
        for i, t in enumerate(tiers):
            ex_name = "MECX" if i == 0 else ""  # 按样表拼写
            lev = to_leverage_str(t.get("mlev"))
            notional = parse_number(t.get("notional_usdt"))
            mmr = to_percent_str(t.get("mmr"))
            rows.append([ex_name, lev, notional, mmr])
        return rows or [["MECX", "", "", ""]]
    elif ex == "weex":
        tiers = sources[ex].get(sym, [])
        tiers = sorted(tiers, key=lambda x: parse_number(x.get("mlev")) or 0)
        rows = []
        for i, t in enumerate(tiers):
            ex_name = "WEEX" if i == 0 else ""
            lev = to_leverage_str(t.get("mlev"))
            rng = t.get("range")
            notional = weex_range_upper(rng) if isinstance(rng, str) else None
            mmr = t.get("mmr") or ""  # weex 已是百分数字符串
            rows.append([ex_name, lev, notional, mmr])
        return rows or [["WEEX", "", "", ""]]
    elif ex == "surf":
        tiers = sources[ex].get(sym, [])
        rows = []
        for i, t in enumerate(tiers):
            ex_name = "SURF" if i == 0 else ""
            lev = to_leverage_str(t.get("mlev"))
            notional = parse_number(t.get("notional_usdt"))
            mmr = to_percent_str(t.get("mmr"))
            rows.append([ex_name, lev, notional, mmr])
        return rows or [["SURF", "", "", ""]]
    else:
        return []


def autosize(ws) -> None:
    for col in range(1, 5):
        letter = get_column_letter(col)
        max_len = 8
        for cell in ws[letter]:
            val = str(cell.value) if cell.value is not None else ""
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[letter].width = max_len + 2


def beautify_sheet(ws) -> None:
    """统一样式：表头底色、细边框，交易所名称列高亮。"""
    thin = Side(style="thin", color="FF999999")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="FFD9D9D9")  # 浅灰
    ex_fill = PatternFill("solid", fgColor="FFEFEFEF")      # 更浅灰用于交易所名称

    max_row = ws.max_row
    max_col = 4

    # 表头样式
    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_all

    # 数据区样式
    for r in range(2, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border_all
            if c == 1:
                # 交易所名称列左对齐，非空行底色+加粗
                cell.alignment = Alignment(horizontal="left", vertical="center")
                if cell.value:
                    cell.fill = ex_fill
                    cell.font = Font(bold=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")


def _build_html(symbols: List[str], html_payload: Dict[str, Dict[str, List[List[Any]]]], book_name: str) -> Path:
    """生成带下拉的静态 HTML，联动展示四所数据。"""
    # 简单样式与脚本（纯原生，不依赖外链）
    exchanges = ["BINANCE", "WEEX", "MECX", "BYBIT", "SURF"]
    data_json = json.dumps(html_payload, ensure_ascii=False)
    symbols_json = json.dumps(symbols, ensure_ascii=False)
    html = f"""
<!doctype html>
<html lang=zh-CN>
<head>
  <meta charset=utf-8>
  <meta name=viewport content="width=device-width, initial-scale=1">
  <title>Leverage & Margin Dashboard - {book_name}</title>
  <style>
    body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial; margin: 24px; }}
    .bar {{ display:flex; align-items:center; gap:12px; margin-bottom:16px; }}
    select {{ padding:6px 10px; font-size:14px; }}
    h2 {{ margin: 16px 0 8px; font-size:16px; }}
    table {{ border-collapse: collapse; width: 680px; margin-bottom: 18px; }}
    th, td {{ border: 1px solid #999; padding: 6px 8px; text-align: center; }}
    th:first-child, td:first-child {{ text-align: left; }}
    thead th {{ background: #eee; font-weight: 700; }}
  </style>
  <script>
    const DATA = {data_json};
    const SYMBOLS = {symbols_json};
    const EXS = {json.dumps(["BINANCE","WEEX","MECX","BYBIT","SURF"])};
    function onSymbolChange() {{
      const sym = document.getElementById('sym').value;
      render(sym);
    }}
    function createTable(rows) {{
      const tbl = document.createElement('table');
      const thead = document.createElement('thead');
      thead.innerHTML = '<tr><th></th><th>最大杠杆</th><th>最大持仓 (USDT)</th><th>维持保证金率</th></tr>';
      tbl.appendChild(thead);
      const tbody = document.createElement('tbody');
      for (const r of rows) {{
        const tr = document.createElement('tr');
        for (const c of r) {{
          const td = document.createElement('td');
          td.textContent = (c===null||c===undefined)?'':c;
          tr.appendChild(td);
        }}
        tbody.appendChild(tr);
      }}
      tbl.appendChild(tbody);
      return tbl;
    }}
    function render(sym) {{
      const root = document.getElementById('root');
      root.innerHTML = '';
      const payload = DATA[sym] || {{}};
      for (const ex of EXS) {{
        const h = document.createElement('h2');
        h.textContent = ex;
        root.appendChild(h);
        const rows = payload[ex] || [["", "", "", ""]];
        root.appendChild(createTable(rows));
      }}
    }}
    window.addEventListener('DOMContentLoaded', () => {{
      const sel = document.getElementById('sym');
      for (const s of SYMBOLS) {{
        const opt = document.createElement('option');
        opt.value = s; opt.textContent = s; sel.appendChild(opt);
      }}
      const init = SYMBOLS.includes('BTCUSDT') ? 'BTCUSDT' : (SYMBOLS[0] || '');
      sel.value = init;
      render(init);
    }});
  </script>
  </head>
  <body>
    <div class="bar">
      <label for="sym">币种：</label>
      <select id="sym" onchange="onSymbolChange()"></select>
    </div>
    <div id="root"></div>
  </body>
</html>
"""
    out = RESULT_HTML_DIR / f"{book_name}.html"
    out.write_text(html, encoding="utf-8")
    return out


def make_excel() -> Path:
    targets = load_targets()
    sources = {
        "binance": load_binance(),
        "bybit": load_bybit(),
        "mexc": load_mexc(),
        "weex": load_weex(),
        "surf": load_surf(),
    }

    wb = Workbook()
    # 删除默认Sheet
    default_ws = wb.active
    wb.remove(default_ws)

    header = ["", "最大杠杆", "最大持仓 (USDT)", "维持保证金率"]

    # 准备 HTML 数据聚合结构：{symbol: {EX: [[ex_name, lev, notional, mmr], ...]}}
    html_payload: Dict[str, Dict[str, List[List[Any]]]] = {}

    for sym in targets:
        # 仅当至少一个交易所存在该币种时才创建Sheet
        if not any(sym in (sources[ex] or {}) for ex in EX_ORDER):
            continue
        ws = wb.create_sheet(title=sym)
        ws.append(header)
        # 样式
        ws["A1"].font = Font(bold=True)
        ws["B1"].font = Font(bold=True)
        ws["C1"].font = Font(bold=True)
        ws["D1"].font = Font(bold=True)
        # HTML 聚合容器
        html_payload[sym] = {}

        for ex in EX_ORDER:
            rows = build_rows_for_exchange(ex, sym, sources)
            for r in rows:
                ws.append(r)
            # 交易所之间留一个空行
            ws.append(["", "", "", ""]) 
            # 写入 HTML 数据（展示时第一列不需要重复的交易所名，保留与 Excel 一致即可）
            # 将空字符串统一保留，前端按空单元显示
            # 显示区块标题使用大写交易所名
            ex_upper = (
                "BINANCE" if ex=="binance" else
                ("BYBIT" if ex=="bybit" else
                ("MECX" if ex=="mexc" else
                ("WEEX" if ex=="weex" else "SURF")))
            )
            html_payload[sym][ex_upper] = rows or [["", "", "", ""]]
        beautify_sheet(ws)
        autosize(ws)

    ts = time.strftime("%Y%m%d_%H%M%S")
    out = RESULT_DIR / f"Leverage&Margin_{ts}.xlsx"
    wb.save(out)
    # 生成 HTML（带下拉联动）
    # 仅使用实际创建了 Sheet 的币种（即 html_payload 的键集合）
    created_symbols = sorted(html_payload.keys())
    _build_html(created_symbols, html_payload, out.stem)
    return out


if __name__ == "__main__":
    # 运行时守卫日志：用于定位外部误创建目录的问题
    import os
    print("[tableMake] CWD:", os.getcwd())
    print("[tableMake] SCRIPT_DIR:", Path(__file__).resolve().parent)
    print("[tableMake] RESULT_DIR:", RESULT_DIR)
    print("[tableMake] Guard: 本脚本输出到 'result/'")

    path = make_excel()
    print(f"已生成: {path}")
