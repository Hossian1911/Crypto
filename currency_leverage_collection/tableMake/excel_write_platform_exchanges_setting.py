from __future__ import annotations

import os
import re
import sys
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import psycopg2
import psycopg2.extras

# === 配置（与 db_write_platform_exchanges_setting.py 同步） ===
PG_HOST = "platformuser.cluster-custom-csteuf9lw8dv.ap-northeast-1.rds.amazonaws.com"
PG_PORT = 5432
PG_DB = "replication_report"
PG_USER = "platform_exchanges_user"
PG_PASS = "Gdafl(j;390HJDL"
TABLE_NAME = "platform_exchanges_setting_min"

BASE_DIR = Path(__file__).resolve().parent.parent
RESULT_DIR = BASE_DIR / "result"

EXS = {"BINANCE", "WEEX", "MECX", "BYBIT", "SURF"}


def _latest_excel(dir_path: Path) -> Optional[Path]:
    if not dir_path.exists():
        return None
    cands = sorted(dir_path.glob("Leverage&Margin_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    return cands[0] if cands else None


def _parse_leverage(v: Any) -> Optional[int]:
    if v is None:
        return None
    s = str(v).strip().upper()
    if not s:
        return None
    s = s.replace("倍", "X")
    m = re.match(r"^([0-9]+)X$", s)
    if m:
        try:
            return int(m.group(1))
        except Exception:
            return None
    try:
        return int(float(s))
    except Exception:
        return None


def _parse_float(v: Any) -> Optional[float]:
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    s = s.replace(",", "")
    if s.endswith("%"):
        try:
            return float(s[:-1]) / 100.0
        except Exception:
            return None
    try:
        return float(s)
    except Exception:
        return None


def _now_utc_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def parse_sheet(ws) -> List[Dict[str, Any]]:
    # 解析每个交易所区块的全部档位，按出现顺序生成 tier_order（从 1 开始）
    rows: List[Dict[str, Any]] = []
    current_ex: Optional[str] = None
    tier_counter: Dict[str, int] = {}

    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx == 1:
            # 跳过标题行
            continue
        ex_cell, lev, size, mmr = None, None, None, None
        if isinstance(row, tuple) and len(row) >= 4:
            ex_cell = row[0]
            lev = row[1]
            size = row[2]
            mmr = row[3]
        else:
            continue

        ex_text = ex_cell.strip().upper() if isinstance(ex_cell, str) else (ex_cell if ex_cell is not None else "")

        # 空行作为分隔：重置 current_ex
        if (not ex_text) and lev is None and size is None and mmr is None:
            current_ex = None
            continue

        # 若第一列是交易所名，则开始新的区块，并重置该交易所计数
        if isinstance(ex_text, str) and ex_text in EXS:
            current_ex = ex_text
            tier_counter[current_ex] = 0
            continue

        # 进入区块后的数据行（第一列通常为空）
        if not current_ex:
            continue

        tier_counter[current_ex] = tier_counter.get(current_ex, 0) + 1
        rows.append({
            "exchange": current_ex,
            "tier_order": tier_counter[current_ex],
            "max_leverage": _parse_leverage(lev),
            "max_size": _parse_float(size),
            "mmr": _parse_float(mmr),
        })

    return rows


def build_records_from_excel(xlsx_path: Path) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    records: List[Dict[str, Any]] = []
    collected_at = _now_utc_iso()

    for sheet_name in wb.sheetnames:
        # sheet 名即 symbol，如 "ETHUSDT"
        sym = sheet_name.strip().upper()
        if not sym.endswith("USDT"):
            continue
        base = sym[:-4]
        quote = "USDT"
        pair = f"{base}/{quote}"

        ws = wb[sheet_name]
        rows = parse_sheet(ws)
        for r in rows:
            rec = {
                "symbol": sym,
                "pair": pair,
                "base": base,
                "quote": quote,
                "exchange": r.get("exchange"),
                "tier_order": int(r.get("tier_order") or 0),
                "pair_id": None,  # Excel 中无此列
                "max_leverage": r.get("max_leverage"),
                "max_size": r.get("max_size"),
                "mmr": r.get("mmr"),
                "source_url": None,  # Excel 中无此列
                "collected_at": collected_at,
            }
            records.append(rec)

    return records


def pg_connect():
    conn = psycopg2.connect(
        host=PG_HOST,
        port=PG_PORT,
        dbname=PG_DB,
        user=PG_USER,
        password=PG_PASS,
        connect_timeout=10,
        sslmode=os.environ.get("PG_SSLMODE", "prefer"),
    )
    return conn


def get_table_columns(conn, table: str) -> List[str]:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT column_name
            FROM information_schema.columns
            WHERE table_schema = 'public' AND table_name = %s
            """,
            (table,),
        )
        return [r[0] for r in cur.fetchall()]


def upsert_records(conn, table: str, records: List[Dict[str, Any]]):
    if not records:
        print("[warn] no records to write.")
        return

    # 方案C：多档位写入
    required_cols = ["symbol", "exchange", "tier_order", "max_leverage", "max_size", "mmr"]
    cols_in_db = set(get_table_columns(conn, table))
    missing = [c for c in required_cols if c not in cols_in_db]
    if missing:
        raise RuntimeError(f"table {table} missing columns: {missing}. 请先执行建表/迁移 SQL")

    # 构造行，空值兜底
    rows = []
    for rec in records:
        rows.append({
            "symbol": rec.get("symbol") or "",
            "exchange": rec.get("exchange") or "",
            "tier_order": int(rec.get("tier_order") or 0),
            "max_leverage": rec.get("max_leverage") if rec.get("max_leverage") is not None else 0,
            "max_size": rec.get("max_size") if rec.get("max_size") is not None else 0.0,
            "mmr": rec.get("mmr") if rec.get("mmr") is not None else 0.0,
        })

    placeholders = ",".join([f"%({c})s" for c in required_cols])
    collist = ",".join(required_cols)

    with conn:
        with conn.cursor() as cur:
            sql = f"""
                INSERT INTO {table} ({collist})
                VALUES ({placeholders})
                ON CONFLICT (symbol, exchange, tier_order) DO UPDATE SET
                  max_leverage=EXCLUDED.max_leverage,
                  max_size=EXCLUDED.max_size,
                  mmr=EXCLUDED.mmr
            """
            psycopg2.extras.execute_batch(cur, sql, rows, page_size=200)

    print(f"[ok] written: {len(rows)} rows to {table}")


def main():
    xlsx = _latest_excel(RESULT_DIR)
    if not xlsx:
        print(f"[err] 未找到 Excel：{RESULT_DIR}/Leverage&Margin_*.xlsx")
        sys.exit(1)
    print(f"[info] 使用最新 Excel: {xlsx.name}")

    records = build_records_from_excel(xlsx)
    print(f"[build] rows from excel: {len(records)}")

    conn = pg_connect()
    try:
        upsert_records(conn, TABLE_NAME, records)
    finally:
        conn.close()


if __name__ == "__main__":
    main()
