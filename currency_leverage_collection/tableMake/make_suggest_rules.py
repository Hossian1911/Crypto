from __future__ import annotations

import json
import time
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data" / "dataGet_api"
HTML_DIR = ROOT / "result" / "html"
OUT_DIR = ROOT / "result" / "suggest"
OUT_DIR.mkdir(parents=True, exist_ok=True)

MAJOR_TIERS = [50_000, 200_000, 500_000, 1_000_000]
MINOR_TIERS = [20_000, 100_000, 200_000]

EXS_SHOW = ["BINANCE", "WEEX", "MECX", "BYBIT"]


def _latest_json() -> Optional[Path]:
    files = sorted(HTML_DIR.glob("Leverage&Margin_*.json"))
    return files[-1] if files else None


def _num(v: Any) -> Optional[float]:
    try:
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).replace(",", "").strip()
        if s == "":
            return None
        return float(s)
    except Exception:
        return None


def _mmr_to_float(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s.endswith('%'):
        try:
            return float(s.rstrip('%')) / 100.0
        except Exception:
            return None
    return _num(s)


def _lev_to_float(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).upper().replace(" ", "").strip()
    if s.endswith("X"):
        s = s[:-1]
    return _num(s)


def _lev_fmt(x: float) -> str:
    if float(x).is_integer():
        return f"{int(x)}X"
    return f"{x}X"


def _im_from_lev(lev: float) -> float:
    return 1.0 / max(1e-9, lev)


def _mmr_fmt(v: float) -> str:
    return f"{v*100:.2f}%"


def _pos_fmt(v: float) -> str:
    if float(v).is_integer():
        return f"{int(v):,}"
    return f"{v:,.2f}"

def _select_tier_for_threshold(rows: List[List[Any]], S: float) -> Optional[Tuple[float, float]]:
    """
    选择“刚好覆盖 S 的那一档”：
    - 取满足 pos >= S 的档位中，pos 最小的那一档（即“比略大的挡位”）。
    - 若所有行 pos < S，则返回 None（该所不覆盖该层级）。
    输入 rows 格式：[[ex_name, lev_str, pos_float, mmr_str], ...]
    返回 (leverage_float, mmr_float)
    """
    parsed: List[Tuple[float, float, float]] = []
    for r in rows:
        if len(r) < 4:
            continue
        lev = _lev_to_float(r[1])
        pos = _num(r[2])
        mmr = _mmr_to_float(r[3])
        if lev is None or pos is None or mmr is None:
            continue
        parsed.append((lev, pos, mmr))
    if not parsed:
        return None
    # 按 pos 从大到小排序，然后线性扫描，选“最后一个仍满足 pos>=S 的行”
    parsed.sort(key=lambda t: t[1], reverse=True)
    chosen: Optional[Tuple[float, float, float]] = None
    for lev, pos, mmr in parsed:
        if pos >= S:
            chosen = (lev, pos, mmr)
        else:
            break
    if chosen is None:
        return None
    return (chosen[0], chosen[2])


def _street_for_symbol(payload: Dict[str, Dict[str, List[List[Any]]]], sym: str, S: float) -> Tuple[Optional[float], Optional[str], Optional[float], Optional[str]]:
    """跨所基准：返回 (max_leverage, max_lev_source, min_mmr, min_mmr_source) for level S。"""
    max_lev_val: Optional[float] = None
    max_lev_src: Optional[str] = None
    min_mmr_val: Optional[float] = None
    min_mmr_src: Optional[str] = None
    sym_map = payload.get(sym) or {}
    for ex in EXS_SHOW:
        rows = sym_map.get(ex) or []
        pick = _select_tier_for_threshold(rows, S)
        if pick is None:
            continue
        lev, mmr = pick
        if lev is not None and (max_lev_val is None or lev > max_lev_val):
            max_lev_val = lev
            max_lev_src = ex
        if mmr is not None and (min_mmr_val is None or mmr < min_mmr_val):
            min_mmr_val = mmr
            min_mmr_src = ex
    return max_lev_val, max_lev_src, min_mmr_val, min_mmr_src


def _build_groups() -> Tuple[List[str], List[str]]:
    # Majors: 从 cmc_top20.json 读取，已在 dataGet_main 里拉取
    cmc_path = DATA_DIR / "cmc" / "cmc_top20.json"
    majors: List[str] = []
    if cmc_path.exists():
        try:
            arr = json.loads(cmc_path.read_text(encoding="utf-8"))
            for it in arr:
                s = str(it.get("name") or "").upper().strip()
                if s:
                    majors.append(f"{s}USDT")
        except Exception:
            pass
    # 所有 SURF 支持的 USDT 交易对
    surf_pairs = (ROOT / "data" / "currency_kinds" / "surf_pairs.json").read_text(encoding="utf-8")
    sp = json.loads(surf_pairs)
    all_syms = []
    for p in sp.get("pairs", []):
        base = str(p.get("base") or "").upper().strip()
        quote = str(p.get("quote") or "").upper().strip()
        if base and quote == "USDT":
            all_syms.append(f"{base}{quote}")
    majors = [s for s in majors if s in set(all_syms)]
    minors = sorted([s for s in set(all_syms) if s not in set(majors)])
    return majors, minors


def generate_excel() -> Path:
    latest = _latest_json()
    if latest is None:
        raise FileNotFoundError("未找到 result/html 下的 Leverage&Margin_*.json")
    data = json.loads(latest.read_text(encoding="utf-8"))
    payload: Dict[str, Dict[str, List[List[Any]]]] = data.get("data") or {}

    majors, minors = _build_groups()

    wb = Workbook()
    # 清理默认sheet
    default_ws = wb.active
    wb.remove(default_ws)

    tiers_json: Dict[str, List[Dict[str, Any]]] = {}

    def write_sheet(sym: str, tiers: List[int]) -> None:
        ws = wb.create_sheet(title=f"{sym} Suggest Rule")
        header = ["max position size", "Max Leverage", "Max Lev Source", "Min MMR", "Min MMR Source", "IM"]
        ws.append(header)
        # 样式
        for c in range(1, len(header) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill("solid", fgColor="FFFFF2AB")
        # 逐层级写入：严格按街上基准聚合
        tiers_json[sym] = []
        for S in tiers:
            street_lev, street_lev_src, street_mmr, street_mmr_src = _street_for_symbol(payload, sym, float(S))
            # 若缺失，留空
            if street_lev is None and street_mmr is None:
                ws.append([_pos_fmt(S), "", "", "", "", ""])
                tiers_json[sym].append({
                    "position": S,
                    "leverage_value": None,
                    "leverage_display": "",
                    "max_lev_source": "",
                    "mmr_value": None,
                    "mmr_display": "",
                    "min_mmr_source": "",
                    "im_value": None,
                    "im_display": "",
                })
                continue
            lev_str = _lev_fmt(street_lev) if street_lev is not None else ""
            mmr_str = _mmr_fmt(street_mmr) if street_mmr is not None else ""
            im_val = (100.0 / float(street_lev)) if street_lev is not None else None
            im_str = f"{im_val:.2f}" if im_val is not None else ""
            ws.append([
                _pos_fmt(S),
                lev_str,
                (street_lev_src or ""),
                mmr_str,
                (street_mmr_src or ""),
                im_str,
            ])
            tiers_json[sym].append({
                "position": S,
                "leverage_value": float(street_lev) if street_lev is not None else None,
                "leverage_display": lev_str,
                "max_lev_source": street_lev_src or "",
                "mmr_value": float(street_mmr) if street_mmr is not None else None,
                "mmr_display": mmr_str,
                "min_mmr_source": street_mmr_src or "",
                "im_value": im_val,
                "im_display": im_str,
            })
        # 边框
        thin = Side(style="thin", color="FFCCCCCC")
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).border = border_all
        # 列宽
        for c in range(1, ws.max_column + 1):
            ws.column_dimensions[chr(ord('A') + c - 1)].width = 18

    for sym in majors:
        write_sheet(sym, MAJOR_TIERS)
    for sym in minors:
        write_sheet(sym, MINOR_TIERS)

    ts = time.strftime('%Y%m%d_%H%M%S')
    out_xlsx = OUT_DIR / f"suggest_rules_{ts}.xlsx"
    wb.save(out_xlsx)

    out_json = OUT_DIR / f"suggest_rules_{ts}.json"
    payload_out = {
        "generated_at": ts,
        "tiers": tiers_json,
    }
    out_json.write_text(json.dumps(payload_out, ensure_ascii=False, indent=2), encoding="utf-8")
    return out_xlsx


if __name__ == "__main__":
    path = generate_excel()
    print(f"Saved: {path}")
